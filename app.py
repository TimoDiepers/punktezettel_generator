import io
import math
import string
from datetime import date

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants / style helpers
# ---------------------------------------------------------------------------
THIN = Side(style="thin")
MEDIUM = Side(style="medium")

FILL_GRAY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
FILL_ORANGE = PatternFill(start_color="F4CCAF", end_color="F4CCAF", fill_type="solid")

FONT_TITLE = Font(bold=True, size=18)
FONT_NORMAL = Font(size=12)
FONT_BOLD_12 = Font(bold=True, size=12)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_CENTER_WRAP = Alignment(
    horizontal="center", vertical="center", wrap_text=True, text_rotation=90
)


# ---------------------------------------------------------------------------
# Session state helpers for dynamic Aufgaben
# ---------------------------------------------------------------------------
def _init_state():
    if "aufgaben" not in st.session_state:
        # Each Aufgabe: list of Teilaufgaben, each Teilaufgabe: {punkte: int, descriptions: list[str]}
        st.session_state.aufgaben = [
            {
                "teilaufgaben": [
                    {"punkte": 4, "descriptions": [""] * 4},
                    {"punkte": 4, "descriptions": [""] * 4},
                ]
            }
        ]


def _add_aufgabe():
    st.session_state.aufgaben.append(
        {
            "teilaufgaben": [
                {"punkte": 4, "descriptions": [""] * 4},
                {"punkte": 4, "descriptions": [""] * 4},
            ]
        }
    )


def _remove_aufgabe(idx: int):
    if len(st.session_state.aufgaben) > 1:
        st.session_state.aufgaben.pop(idx)


def _add_teilaufgabe(a_idx: int):
    st.session_state.aufgaben[a_idx]["teilaufgaben"].append(
        {"punkte": 4, "descriptions": [""] * 4}
    )


def _remove_teilaufgabe(a_idx: int, t_idx: int):
    ta_list = st.session_state.aufgaben[a_idx]["teilaufgaben"]
    if len(ta_list) > 1:
        ta_list.pop(t_idx)


def _sync_descriptions(a_idx: int, t_idx: int, new_punkte: int):
    """Keep descriptions list in sync with punkte count."""
    ta = st.session_state.aufgaben[a_idx]["teilaufgaben"][t_idx]
    old = ta["descriptions"]
    if new_punkte > len(old):
        ta["descriptions"] = old + [""] * (new_punkte - len(old))
    elif new_punkte < len(old):
        ta["descriptions"] = old[:new_punkte]
    ta["punkte"] = new_punkte


# ---------------------------------------------------------------------------
# Dummy template generation
# ---------------------------------------------------------------------------
def generate_template() -> bytes:
    dummy = pd.DataFrame(
        {
            "Matr-Nr": [f"{100000 + i}" for i in range(12)],
            "Nachname": [
                "Müller", "Schmidt", "Schneider", "Fischer", "Weber",
                "Meyer", "Wagner", "Becker", "Schulz", "Hoffmann",
                "Koch", "Richter",
            ],
            "Vorname": [
                "Anna", "Ben", "Clara", "David", "Eva",
                "Felix", "Greta", "Hans", "Ida", "Jan",
                "Klara", "Lukas",
            ],
        }
    )
    buf = io.BytesIO()
    dummy.to_excel(buf, index=False, engine="openpyxl")
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Excel generation helpers
# ---------------------------------------------------------------------------
def _set_medium_outline(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            left = MEDIUM if c == min_col else THIN
            right = MEDIUM if c == max_col else THIN
            top = MEDIUM if r == min_row else THIN
            bottom = MEDIUM if r == max_row else THIN
            ws.cell(row=r, column=c).border = Border(
                left=left, right=right, top=top, bottom=bottom
            )


def build_studenten_sheet(wb: Workbook, students: pd.DataFrame, studis_pro_mappe: int):
    ws = wb.active
    ws.title = "Studenten"

    headers = ["Mappe", "Stelle in Mappe", "KlausurCode", "Matrikelnummer", "Nachname", "Vorname"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.alignment = ALIGN_CENTER

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20

    for idx, (_, row) in enumerate(students.iterrows()):
        r = idx + 2
        mappe = idx // studis_pro_mappe + 1
        stelle = idx % studis_pro_mappe
        ws.cell(row=r, column=1, value=mappe).alignment = ALIGN_CENTER
        ws.cell(row=r, column=2, value=stelle).alignment = ALIGN_CENTER
        ws.cell(row=r, column=3, value=f"{mappe}_{stelle}").alignment = ALIGN_CENTER
        ws.cell(row=r, column=4, value=str(row.iloc[0])).alignment = ALIGN_CENTER
        ws.cell(row=r, column=5, value=str(row.iloc[1])).alignment = ALIGN_CENTER
        ws.cell(row=r, column=6, value=str(row.iloc[2])).alignment = ALIGN_CENTER


def build_mappe_sheet(
    wb: Workbook,
    sheet_name: str,
    aufgaben: list[dict],
    students_in_mappe: pd.DataFrame,
    mappe_nr: int,
    semester: str,
    exam_date,
):
    """Create one sheet per Mappe with all Aufgaben side by side.

    Layout:
        A-E: student info
        Then ALL SUM blocks together: [SS S S | SS S | ...]  (one per Aufgabe)
        Then ALL point columns:       [a b c | a b | a b c d | ...]
    """
    ws = wb.create_sheet(title=sheet_name)
    n_students = len(students_in_mappe)
    n_aufgaben = len(aufgaben)

    # --- Compute column positions ---
    # Phase 1: Place all SUM blocks (SS + S cols) starting at column F
    aufgabe_layouts: list[dict] = []
    col = 6

    for aufgabe in aufgaben:
        teil_punkte = [ta["punkte"] for ta in aufgabe["teilaufgaben"]]
        n_teil = len(teil_punkte)
        ss_col = col
        s_cols = [col + 1 + t for t in range(n_teil)]
        sum_block_end = col + n_teil
        aufgabe_layouts.append({
            "teil_punkte": teil_punkte,
            "teil_descriptions": [ta["descriptions"] for ta in aufgabe["teilaufgaben"]],
            "n_teil": n_teil,
            "ss_col": ss_col,
            "s_cols": s_cols,
            "sum_block_end": sum_block_end,
            "teil_point_starts": [],  # filled in phase 2
        })
        col = sum_block_end + 1

    # Phase 2: Place all point columns after all SUM blocks
    for lay in aufgabe_layouts:
        for pts in lay["teil_punkte"]:
            lay["teil_point_starts"].append(col)
            col += pts

    # --- Row 1: Title + SUM headers ---
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=5)
    ws.cell(row=1, column=1, value=semester).font = FONT_TITLE
    ws.cell(row=1, column=1).alignment = ALIGN_CENTER

    for a_idx, lay in enumerate(aufgabe_layouts):
        aufgabe_nr = a_idx + 1
        # "SUM" merged over SS + S cols, rows 1-2
        ws.merge_cells(start_row=1, end_row=2, start_column=lay["ss_col"], end_column=lay["sum_block_end"])
        ws.cell(row=1, column=lay["ss_col"], value="SUM").font = FONT_TITLE
        ws.cell(row=1, column=lay["ss_col"]).alignment = ALIGN_CENTER

    # --- Row 2: Datum, Mappe, Teilaufgabe labels ---
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=2)
    ws.cell(row=2, column=1, value="Datum:").font = FONT_TITLE
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=2, column=3, value=exam_date).font = FONT_TITLE
    ws.cell(row=2, column=3).alignment = ALIGN_CENTER
    ws.cell(row=2, column=3).number_format = "DD.MM.YYYY"
    ws.cell(row=2, column=4, value="Mappe").font = FONT_TITLE
    ws.cell(row=2, column=4).alignment = ALIGN_CENTER
    ws.cell(row=2, column=5, value=mappe_nr).font = FONT_TITLE
    ws.cell(row=2, column=5).alignment = ALIGN_CENTER

    for a_idx, lay in enumerate(aufgabe_layouts):
        aufgabe_nr = a_idx + 1
        for t_idx, pts in enumerate(lay["teil_punkte"]):
            start_c = lay["teil_point_starts"][t_idx]
            end_c = start_c + pts - 1
            label = f"A{aufgabe_nr}.{t_idx + 1}"
            if end_c > start_c:
                ws.merge_cells(start_row=2, end_row=2, start_column=start_c, end_column=end_c)
            ws.cell(row=2, column=start_c, value=label).font = FONT_TITLE
            ws.cell(row=2, column=start_c).alignment = ALIGN_CENTER

    # --- Row 3: Sub-task descriptions (rotated 90°) ---
    ws.cell(row=3, column=5, value="Richtige Klausur?").font = FONT_NORMAL
    ws.cell(row=3, column=5).alignment = ALIGN_CENTER_WRAP
    ws.cell(row=3, column=5).fill = FILL_GRAY

    for a_idx, lay in enumerate(aufgabe_layouts):
        aufgabe_nr = a_idx + 1
        n_teil = lay["n_teil"]

        # SS description
        label_text = " + ".join(f"A{aufgabe_nr}.{t + 1}" for t in range(n_teil))
        ws.cell(row=3, column=lay["ss_col"], value=label_text).font = FONT_NORMAL
        ws.cell(row=3, column=lay["ss_col"]).alignment = ALIGN_CENTER_WRAP
        ws.cell(row=3, column=lay["ss_col"]).fill = FILL_GRAY

        # S descriptions
        for t_idx in range(n_teil):
            ws.cell(row=3, column=lay["s_cols"][t_idx], value=f"A{aufgabe_nr}.{t_idx + 1}").font = FONT_NORMAL
            ws.cell(row=3, column=lay["s_cols"][t_idx]).alignment = ALIGN_CENTER_WRAP
            ws.cell(row=3, column=lay["s_cols"][t_idx]).fill = FILL_GRAY

        # Point column descriptions
        descs_list = lay["teil_descriptions"]
        for t_idx in range(n_teil):
            descs = descs_list[t_idx] if t_idx < len(descs_list) else []
            for p in range(lay["teil_punkte"][t_idx]):
                pc = lay["teil_point_starts"][t_idx] + p
                desc = descs[p] if p < len(descs) else ""
                if desc:
                    ws.cell(row=3, column=pc, value=desc).font = FONT_NORMAL
                ws.cell(row=3, column=pc).alignment = ALIGN_CENTER_WRAP
                ws.cell(row=3, column=pc).fill = FILL_GRAY

    ws.row_dimensions[3].height = 148

    # --- Row 4: Column headers ---
    for c, v in {1: "Nr", 2: "Matr-Nr", 3: "Nachname", 4: "Vorname"}.items():
        cell = ws.cell(row=4, column=c, value=v)
        cell.font = FONT_BOLD_12
        cell.alignment = ALIGN_CENTER
        cell.fill = FILL_GRAY
    ws.cell(row=4, column=5).fill = FILL_GRAY
    ws.cell(row=4, column=5).alignment = ALIGN_CENTER

    letters = list(string.ascii_lowercase)
    for lay in aufgabe_layouts:
        ws.cell(row=4, column=lay["ss_col"], value="SS").font = FONT_BOLD_12
        ws.cell(row=4, column=lay["ss_col"]).alignment = ALIGN_CENTER
        ws.cell(row=4, column=lay["ss_col"]).fill = FILL_GRAY

        for t_idx in range(lay["n_teil"]):
            ws.cell(row=4, column=lay["s_cols"][t_idx], value="S").font = FONT_BOLD_12
            ws.cell(row=4, column=lay["s_cols"][t_idx]).alignment = ALIGN_CENTER
            ws.cell(row=4, column=lay["s_cols"][t_idx]).fill = FILL_GRAY

            for p in range(lay["teil_punkte"][t_idx]):
                pc = lay["teil_point_starts"][t_idx] + p
                ws.cell(row=4, column=pc, value=letters[p % 26]).font = FONT_BOLD_12
                ws.cell(row=4, column=pc).alignment = ALIGN_CENTER
                ws.cell(row=4, column=pc).fill = FILL_ORANGE

    ws.row_dimensions[4].height = 21

    # --- Row 5: Maximale Punkte ---
    ws.merge_cells(start_row=5, end_row=5, start_column=1, end_column=5)
    ws.cell(row=5, column=1, value="Maximale Punkte").font = FONT_BOLD_12
    ws.cell(row=5, column=1).alignment = ALIGN_CENTER
    for c in range(1, 6):
        ws.cell(row=5, column=c).fill = FILL_GRAY

    for lay in aufgabe_layouts:
        # SS = SUM of S cols
        ws.cell(row=5, column=lay["ss_col"],
                value=f"=SUM({get_column_letter(lay['s_cols'][0])}5:{get_column_letter(lay['s_cols'][-1])}5)").font = FONT_NORMAL
        ws.cell(row=5, column=lay["ss_col"]).alignment = ALIGN_CENTER
        ws.cell(row=5, column=lay["ss_col"]).fill = FILL_GRAY

        for t_idx in range(lay["n_teil"]):
            fp = get_column_letter(lay["teil_point_starts"][t_idx])
            lp = get_column_letter(lay["teil_point_starts"][t_idx] + lay["teil_punkte"][t_idx] - 1)
            ws.cell(row=5, column=lay["s_cols"][t_idx], value=f"=SUM({fp}5:{lp}5)").font = FONT_NORMAL
            ws.cell(row=5, column=lay["s_cols"][t_idx]).alignment = ALIGN_CENTER
            ws.cell(row=5, column=lay["s_cols"][t_idx]).fill = FILL_GRAY

            for p in range(lay["teil_punkte"][t_idx]):
                pc = lay["teil_point_starts"][t_idx] + p
                ws.cell(row=5, column=pc, value=1).font = FONT_NORMAL
                ws.cell(row=5, column=pc).alignment = ALIGN_CENTER
                ws.cell(row=5, column=pc).fill = FILL_GRAY

    # --- Row 6: Durchschnittliche Punkte ---
    ws.merge_cells(start_row=6, end_row=6, start_column=1, end_column=5)
    ws.cell(row=6, column=1, value="Durchschnittliche Punkte").font = FONT_BOLD_12
    ws.cell(row=6, column=1).alignment = ALIGN_CENTER
    for c in range(1, 6):
        ws.cell(row=6, column=c).fill = FILL_GRAY

    first_data_row = 7
    last_data_row = 7 + n_students - 1

    for lay in aufgabe_layouts:
        ws.cell(row=6, column=lay["ss_col"],
                value=f"=SUM({get_column_letter(lay['s_cols'][0])}6:{get_column_letter(lay['s_cols'][-1])}6)").font = FONT_NORMAL
        ws.cell(row=6, column=lay["ss_col"]).alignment = ALIGN_CENTER
        ws.cell(row=6, column=lay["ss_col"]).fill = FILL_GRAY
        ws.cell(row=6, column=lay["ss_col"]).number_format = "0.00"

        for t_idx in range(lay["n_teil"]):
            fp_l = get_column_letter(lay["teil_point_starts"][t_idx])
            lp_l = get_column_letter(lay["teil_point_starts"][t_idx] + lay["teil_punkte"][t_idx] - 1)
            ws.cell(row=6, column=lay["s_cols"][t_idx], value=f"=SUM({fp_l}6:{lp_l}6)").font = FONT_NORMAL
            ws.cell(row=6, column=lay["s_cols"][t_idx]).alignment = ALIGN_CENTER
            ws.cell(row=6, column=lay["s_cols"][t_idx]).fill = FILL_GRAY
            ws.cell(row=6, column=lay["s_cols"][t_idx]).number_format = "0.00"

            for p in range(lay["teil_punkte"][t_idx]):
                pc = lay["teil_point_starts"][t_idx] + p
                pc_l = get_column_letter(pc)
                ws.cell(
                    row=6, column=pc,
                    value=f'=IFERROR(AVERAGE({pc_l}{first_data_row}:{pc_l}{last_data_row}),"")',
                ).font = FONT_NORMAL
                ws.cell(row=6, column=pc).alignment = ALIGN_CENTER
                ws.cell(row=6, column=pc).fill = FILL_GRAY
                ws.cell(row=6, column=pc).number_format = "0.00"

    # --- Student rows ---
    for s_idx, (_, student) in enumerate(students_in_mappe.iterrows()):
        r = 7 + s_idx
        ws.cell(row=r, column=1, value=s_idx).alignment = ALIGN_CENTER
        ws.cell(row=r, column=2, value=str(student.iloc[0])).alignment = ALIGN_CENTER
        ws.cell(row=r, column=2).font = FONT_NORMAL
        ws.cell(row=r, column=3, value=str(student.iloc[1])).alignment = ALIGN_CENTER
        ws.cell(row=r, column=3).font = FONT_NORMAL
        ws.cell(row=r, column=4, value=str(student.iloc[2])).alignment = ALIGN_CENTER
        ws.cell(row=r, column=4).font = FONT_NORMAL

        for lay in aufgabe_layouts:
            # SS = SUM of S cols
            ws.cell(row=r, column=lay["ss_col"],
                    value=f"=SUM({get_column_letter(lay['s_cols'][0])}{r}:{get_column_letter(lay['s_cols'][-1])}{r})").font = FONT_NORMAL
            ws.cell(row=r, column=lay["ss_col"]).alignment = ALIGN_CENTER
            ws.cell(row=r, column=lay["ss_col"]).fill = FILL_GRAY

            for t_idx in range(lay["n_teil"]):
                fp = get_column_letter(lay["teil_point_starts"][t_idx])
                lp = get_column_letter(lay["teil_point_starts"][t_idx] + lay["teil_punkte"][t_idx] - 1)
                ws.cell(row=r, column=lay["s_cols"][t_idx], value=f"=SUM({fp}{r}:{lp}{r})").font = FONT_NORMAL
                ws.cell(row=r, column=lay["s_cols"][t_idx]).alignment = ALIGN_CENTER
                ws.cell(row=r, column=lay["s_cols"][t_idx]).fill = FILL_GRAY

        ws.row_dimensions[r].height = 16

    # --- Column widths ---
    ws.column_dimensions["A"].width = 5.5
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 8
    for lay in aufgabe_layouts:
        ws.column_dimensions[get_column_letter(lay["ss_col"])].width = 5
        for t_idx in range(lay["n_teil"]):
            ws.column_dimensions[get_column_letter(lay["s_cols"][t_idx])].width = 4.5
            for p in range(lay["teil_punkte"][t_idx]):
                pc = lay["teil_point_starts"][t_idx] + p
                ws.column_dimensions[get_column_letter(pc)].width = 4.5

    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[5].height = 16
    ws.row_dimensions[6].height = 17

    # --- Borders ---
    last_data_row_border = 6 + n_students
    _set_medium_outline(ws, 4, last_data_row_border, 1, 5)
    for lay in aufgabe_layouts:
        # SUM block
        _set_medium_outline(ws, 3, last_data_row_border, lay["ss_col"], lay["sum_block_end"])
        # Each Teilaufgabe point block
        for t_idx in range(lay["n_teil"]):
            start_c = lay["teil_point_starts"][t_idx]
            end_c = start_c + lay["teil_punkte"][t_idx] - 1
            _set_medium_outline(ws, 3, last_data_row_border, start_c, end_c)


def generate_excel(
    students: pd.DataFrame,
    aufgaben: list[dict],
    studis_pro_mappe: int,
    semester: str,
    exam_date,
) -> bytes:
    wb = Workbook()
    n_students = len(students)
    n_mappen = math.ceil(n_students / studis_pro_mappe)

    build_studenten_sheet(wb, students, studis_pro_mappe)

    for m in range(n_mappen):
        mappe_nr = m + 1
        start_idx = m * studis_pro_mappe
        end_idx = min(start_idx + studis_pro_mappe, n_students)
        students_in_mappe = students.iloc[start_idx:end_idx]

        sheet_name = f"Mappe {mappe_nr}"

        build_mappe_sheet(
            wb=wb,
            sheet_name=sheet_name,
            aufgaben=aufgaben,
            students_in_mappe=students_in_mappe,
            mappe_nr=mappe_nr,
            semester=semester,
            exam_date=exam_date,
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Streamlit UI
# ---------------------------------------------------------------------------
st.set_page_config(page_title="Punktezettel Generator", layout="wide")
st.title("Punktezettel Generator")

_init_state()

# --- Template download + file upload side by side ---
col_upload, col_template = st.columns([3, 1])
with col_upload:
    uploaded_file = st.file_uploader(
        "Studierendenliste hochladen (Excel: Matr-Nr, Nachname, Vorname)",
        type=["xlsx", "xls"],
    )
with col_template:
    st.markdown("<br>", unsafe_allow_html=True)
    st.download_button(
        label="Vorlage herunterladen",
        data=generate_template(),
        file_name="Studierendenliste_Vorlage.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# --- General settings ---
col1, col2 = st.columns(2)
with col1:
    semester = st.text_input("Semester", value="WiSe 25/26")
    exam_date = st.date_input("Klausurdatum", value=date.today())
with col2:
    studis_pro_mappe = st.number_input("Studis pro Mappe", min_value=1, value=5)

# --- Aufgaben configuration ---
st.subheader("Aufgaben konfigurieren")

for a_idx, aufgabe in enumerate(st.session_state.aufgaben):
    with st.expander(f"Aufgabe {a_idx + 1}", expanded=(a_idx == 0)):
        # Remove Aufgabe button
        if len(st.session_state.aufgaben) > 1:
            if st.button(
                f"Aufgabe {a_idx + 1} entfernen",
                key=f"rm_aufgabe_{a_idx}",
                type="secondary",
            ):
                _remove_aufgabe(a_idx)
                st.rerun()

        for t_idx, ta in enumerate(aufgabe["teilaufgaben"]):
            st.markdown(f"**Teilaufgabe {a_idx + 1}.{t_idx + 1}**")

            tcol1, tcol2 = st.columns([1, 3])
            with tcol1:
                new_punkte = st.number_input(
                    f"Punkte A{a_idx + 1}.{t_idx + 1}",
                    min_value=1,
                    max_value=50,
                    value=ta["punkte"],
                    key=f"punkte_{a_idx}_{t_idx}",
                )
                if new_punkte != ta["punkte"]:
                    _sync_descriptions(a_idx, t_idx, int(new_punkte))
                    st.rerun()

            with tcol2:
                # Optional descriptions for each point column
                desc_cols = st.columns(min(ta["punkte"], 6))
                letters = list(string.ascii_lowercase)
                for p_idx in range(ta["punkte"]):
                    with desc_cols[p_idx % len(desc_cols)]:
                        new_desc = st.text_input(
                            f"{letters[p_idx % 26]})",
                            value=ta["descriptions"][p_idx] if p_idx < len(ta["descriptions"]) else "",
                            key=f"desc_{a_idx}_{t_idx}_{p_idx}",
                            placeholder="Beschreibung (optional)",
                        )
                        if p_idx < len(ta["descriptions"]):
                            ta["descriptions"][p_idx] = new_desc

            # Remove Teilaufgabe
            if len(aufgabe["teilaufgaben"]) > 1:
                if st.button(
                    f"Teilaufgabe {a_idx + 1}.{t_idx + 1} entfernen",
                    key=f"rm_teil_{a_idx}_{t_idx}",
                ):
                    _remove_teilaufgabe(a_idx, t_idx)
                    st.rerun()

            st.divider()

        if st.button(f"Teilaufgabe hinzufügen", key=f"add_teil_{a_idx}"):
            _add_teilaufgabe(a_idx)
            st.rerun()

if st.button("Aufgabe hinzufügen", type="secondary"):
    _add_aufgabe()
    st.rerun()

# ---------------------------------------------------------------------------
# Generate & Download
# ---------------------------------------------------------------------------
st.divider()

if uploaded_file is not None:
    try:
        df = pd.read_excel(uploaded_file)
    except Exception as e:
        st.error(f"Fehler beim Lesen der Datei: {e}")
        st.stop()

    if len(df.columns) < 3:
        st.error("Die Datei muss mindestens 3 Spalten haben (Matr-Nr, Nachname, Vorname).")
        st.stop()

    students = df.iloc[:, :3].copy()
    students.columns = ["Matr-Nr", "Nachname", "Vorname"]
    students = students.dropna(how="all")

    n_mappen = math.ceil(len(students) / studis_pro_mappe)

    st.success(
        f"{len(students)} Studierende geladen → {n_mappen} Mappen "
        f"(à {studis_pro_mappe}, letzte Mappe: {len(students) - (n_mappen - 1) * studis_pro_mappe})"
    )

    st.dataframe(students, use_container_width=True, hide_index=True)

    n_aufgaben = len(st.session_state.aufgaben)
    total_sheets = n_mappen + 1
    total_points = sum(
        sum(ta["punkte"] for ta in a["teilaufgaben"])
        for a in st.session_state.aufgaben
    )
    st.info(
        f"Es werden **{total_sheets} Sheets** erstellt "
        f"({n_mappen} Mappen + Studenten-Übersicht, {n_aufgaben} Aufgaben pro Blatt). "
        f"Gesamtpunkte: **{total_points}**"
    )

    if st.button("Punktezettel erstellen", type="primary"):
        with st.spinner("Excel wird erstellt..."):
            excel_bytes = generate_excel(
                students=students,
                aufgaben=st.session_state.aufgaben,
                studis_pro_mappe=int(studis_pro_mappe),
                semester=semester,
                exam_date=exam_date,
            )

        st.download_button(
            label="Download Punktezettel (.xlsx)",
            data=excel_bytes,
            file_name=f"Punktezettel_{semester.replace(' ', '_').replace('/', '')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
else:
    st.info("Bitte eine Studierendenliste hochladen, um zu beginnen.")
